import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, colors, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
from os.path import exists
import os
from os import listdir
import pprint


readfile = 'UMAP_PWR2 V1.0.xlsx'
writefile = 'BOM Default Template1.xlsx'
column_titles = ['Adress', 'LibRef', 'Description', 'Designator', 'Quantity']

if __name__ == '__main__':
    wb = load_workbook(readfile)
    ws = wb.active
    wb1 = load_workbook(writefile)
    ws1 = wb1.active

    for k in range(1, len(column_titles) + 1):
        for m in range(0, len(column_titles)):
            # ws1.cell(row=11, column=k, value=column_titles[m])
            ws1.append({11:123, 12:111111})

    # for row in range(2, ws.max_row + 1):
    #     # Each row in the spreadsheet has data for one census tract.
    #     column1 = ws['A' + str(row)].value
    #     column2 = ws['B' + str(row)].value
    #     column3 = ws['C' + str(row)].value
    #     column4 = ws['D' + str(row)].value
    #     column5 = ws['E' + str(row)].value
        # print([column1])
        # print([column2])
        # print(ws.max_row)

        # for i in range(12, ws.max_row):
        #     for j in range(1, len(column_titles)+1)
        #         ws1.cell(row = i,column=j [column1])


    # wb.save(writefile)
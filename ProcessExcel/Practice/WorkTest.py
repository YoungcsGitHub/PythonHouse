import openpyxl

if __name__ == '__main__':
    # 生成一个新的空的Excel
    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = 'WorkSheetTitle'

    ws2 = wb.create_sheet('NewWorkSheet2')
    ws3 = wb.create_sheet('NewWorkSheet3', 0)

    ws.sheet_properties.tabColor = '1072BA'

    ws['A1'] = 'HOGE'
    ws['B1'] = 'FUGA'
    ws.cell(row= 4, column= 2, value=10)
    v = 0
    for i in range(5,16):
        for n in range(1,10):
            ws.cell(row= i, column= n, value= v)
            v += 1

    # 保存
    wb.save('example.xlsx')

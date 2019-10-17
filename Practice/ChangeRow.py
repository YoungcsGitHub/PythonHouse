import openpyxl

if __name__ == '__main__':
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = 'WorkSheetTitle'

    # 单元格内换行
    ws['A1'] = "A\nB\nC"
    ws['A1'].alignment = openpyxl.styles.Alignment(wrapText = True)

    wb.save('example2.xlsx')